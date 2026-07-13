"""数据导出服务"""
from __future__ import annotations

import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def export_to_csv(columns: list[str], rows: list[dict]) -> bytes:
    """将查询结果导出为 CSV 字节流"""
    output: io.StringIO = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([row.get(col, '') for col in columns])
    return output.getvalue().encode('utf-8-sig')


def export_to_excel(columns: list[str], rows: list[dict], sheet_name: str = "查询结果") -> bytes:
    """将查询结果导出为格式化的 Excel 文件"""
    wb: openpyxl.Workbook = openpyxl.Workbook()
    ws: openpyxl.worksheet.worksheet.Worksheet = wb.active
    ws.title = sheet_name

    # 表头样式
    header_fill: PatternFill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font: Font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment: Alignment = Alignment(horizontal="center", vertical="center")
    thin_border: Border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # 写入表头
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    data_font: Font = Font(size=10)
    data_alignment: Alignment = Alignment(vertical="center")
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    # 自动调整列宽
    for col_idx, col_name in enumerate(columns, 1):
        max_length: int = len(str(col_name))
        for row in rows:
            val: str = str(row.get(col_name, ''))
            max_length = max(max_length, len(val))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 4, 40)

    # 冻结首行
    ws.freeze_panes = "A2"

    buffer: io.BytesIO = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
